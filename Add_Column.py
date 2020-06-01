import openpyxl as xl
from datetime import datetime
import xlsxwriter


def create_copy(src_file, dest_file):
    workbook = xlsxwriter.Workbook(dest_file)
    workbook.close()

    ################## FOR TAB 1 ##########################

    wb1 = xl.load_workbook(filename=src_file)
    ws1 = wb1.worksheets[0]

    wb2 = xl.load_workbook(filename=dest_file)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(dest_file)

    ################## FOR TAB 2 ##########################

    wb1 = xl.load_workbook(filename=src_file)
    ws1 = wb1.worksheets[1]

    wb2 = xl.load_workbook(filename=dest_file)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(dest_file)

    ################## FOR TAB 3 ##########################

    wb1 = xl.load_workbook(filename=src_file)
    ws1 = wb1.worksheets[2]

    wb3 = xl.load_workbook(filename=dest_file)
    ws3 = wb3.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws3[cell.coordinate].value = cell.value

    wb3.save(dest_file)


def create_column(data, file_loc, year, month):
    ################## FOR TAB 1 ##########################

    wb1 = xl.load_workbook(filename=file_loc)
    ws1 = wb1.worksheets[1]

    row_count = ws1.max_row

    # Inserting a column into the worksheet at 2nd position
    ws1.insert_cols(0)
    ws1.insert_cols(11)
    ws1.insert_cols(12)
    ws1.insert_cols(13)
    ws1.cell(row=1, column=1).value = "Seller UID"
    ws1.cell(row=1, column=11).value = "Insert Date"
    ws1.cell(row=1, column=12).value = "Insert Month"
    ws1.cell(row=1, column=13).value = "Insert Year"
    #
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    for row in range(2, row_count+1):
        ws1.cell(row, 1).value = data
        ws1.cell(row, 11).value = month
        ws1.cell(row, 12).value = year
        ws1.cell(row, 13).value = date

    wb1.save(file_loc)

    ################## FOR TAB 2 ##########################

    wb2 = xl.load_workbook(filename=file_loc)
    ws2 = wb2.worksheets[2]
    row_count = ws2.max_row

    # Inserting a column into the worksheet at 2nd position
    ws2.insert_cols(0)
    ws2.insert_cols(17)
    ws2.insert_cols(18)
    ws2.insert_cols(19)
    ws2.cell(row=1, column=1).value = "Seller UID"
    ws2.cell(row=1, column=17).value = "Insert Date"
    ws2.cell(row=1, column=18).value = "Insert Month"
    ws2.cell(row=1, column=19).value = "Insert Year"
    #
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    for row in range(2, row_count+1):
        ws2.cell(row, 1).value = data
        ws2.cell(row, 17).value = month
        ws2.cell(row, 18).value = year
        ws2.cell(row, 19).value = date

    wb2.save(file_loc)

    ################## FOR TAB 3 ##########################

    wb3 = xl.load_workbook(filename=file_loc)
    ws3 = wb3.worksheets[3]
    row_count = ws3.max_row

    # Inserting a column into the worksheet at 2nd position
    ws3.insert_cols(0)
    ws3.insert_cols(8)
    ws3.insert_cols(9)
    ws3.insert_cols(10)
    ws3.cell(row=1, column=1).value = "Seller UID"
    ws3.cell(row=1, column=8).value = "Insert Date"
    ws3.cell(row=1, column=9).value = "Insert Month"
    ws3.cell(row=1, column=10).value = "Insert Year"
    #
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    for row in range(2, row_count+1):
        ws3.cell(row, 1).value = data
        ws3.cell(row, 8).value = month
        ws3.cell(row, 9).value = year
        ws3.cell(row, 10).value = date

    wb3.save(file_loc)
